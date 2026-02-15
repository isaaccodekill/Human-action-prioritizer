import json
from pathlib import Path

from openpyxl import load_workbook

from prioritizer.models import Action

PACKAGE_DIR = Path(__file__).resolve().parent.parent
XLSX_PATH = PACKAGE_DIR / "data" / "actions.xlsx"
JSON_PATH = PACKAGE_DIR / "data" / "actions.json"


# Map CSV headers to model field names
HEADER_MAP = {
    "Action": "action",
    "Solution": "solution",
    "Solution Classification": "solution_classification",
    "Mode": "mode",
    "Sector": "sector",
    "Cluster": "cluster",
    "Adoption Unit": "adoption_unit",
    "Effectiveness t CO₂‑eq (100-yr)/unit": "effectiveness",
    "Adoption Current": "adoption_current",
    "Adoption Achievable Range": "adoption_achievable_range",
    "GHG Impact Gt CO₂-eq (100‑yr)/yr": "ghg_impact",
    "Cost US$ per t CO₂‑eq": "cost",
    "Climate Pollutants Mitigated": "climate_pollutants_mitigated",
    "Speed of Action": "speed_of_action",
    "Climate Adaptation Benefits": "climate_adaptation_benefits",
    "Environment Benefits": "environment_benefits",
    "Human Well-being Benefits": "human_wellbeing_benefits",
}


def load_actions_from_excel(file_path: Path) -> list[Action]:
    actions = []
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Row 1 is a title row, row 2 has the actual headers
    raw_headers = [cell.value for cell in sheet[2]]
    # Normalize non-breaking spaces to regular spaces
    headers = [h.replace("\xa0", " ") if h else h for h in raw_headers]
    for row in sheet.iter_rows(min_row=3, values_only=True):
        mapped = {}
        for xlsx_key, model_key in HEADER_MAP.items():
            val = row[headers.index(xlsx_key)]
            if val is None:
                mapped[model_key] = None
            elif isinstance(val, str):
                mapped[model_key] = val.strip() or None
            else:
                mapped[model_key] = str(val)
        actions.append(Action(**mapped))
    return actions

def save_actions_to_json(actions: list[Action], file_path: Path):
    with open(file_path, mode="w", encoding="utf-8") as jsonfile:
        json.dump([action.model_dump() for action in actions], jsonfile, indent=4)


if __name__ == "__main__":
    all_actions = load_actions_from_excel(XLSX_PATH)
    tasks = [a for a in all_actions if a.ghg_impact is not None]
    save_actions_to_json(tasks, JSON_PATH)
    print(f"Loaded {len(all_actions)} actions, {len(tasks)} with GHG impact data")
    print(f"Saved to {JSON_PATH}")